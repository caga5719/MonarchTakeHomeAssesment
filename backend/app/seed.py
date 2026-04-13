"""
Seed reference tables (gl_codes, properties) from xlsx files.

Usage:
    uv run python app/seed.py
"""

from pathlib import Path
import openpyxl

import table_models  # noqa: F401 — registers SQLModel tables with metadata
from database import engine, init_db
from sqlmodel import Session
from table_models import GLCode, Property

REPO_ROOT = Path(__file__).resolve().parents[2]
GL_LIST_PATH = REPO_ROOT / "GL List.xlsx"
PROPERTY_LIST_PATH = REPO_ROOT / "Property List.xlsx"


def seed_gl_codes(session: Session) -> int:
    """Load all GL codes from the xlsx and upsert into gl_codes. Returns row count."""
    wb = openpyxl.load_workbook(GL_LIST_PATH, read_only=True, data_only=True)
    ws = wb.active  # only one sheet: 'ySQL_0_10042026110022 (1)'

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    data = [(int(scode), str(sdesc).strip()) for scode, sdesc in rows if scode is not None and sdesc is not None]

    for scode, sdesc in data:
        session.merge(GLCode(scode=scode, sdesc=sdesc))

    session.commit()
    wb.close()
    return len(data)


def seed_properties(session: Session) -> int:
    """
    Load currently-owned properties, then enrich with metadata from Other Properties
    where the yardi_code matches. All yardi_codes are normalized to UPPERCASE.
    Returns total row count inserted.
    """
    wb = openpyxl.load_workbook(PROPERTY_LIST_PATH, read_only=True, data_only=True)

    # --- Currently Owned: columns are Website ID, Yardi Code ---
    ws_owned = wb["Currently Owned"]
    owned_rows = list(ws_owned.iter_rows(min_row=2, values_only=True))
    owned: dict[str, dict] = {}
    for row in owned_rows:
        website_id = row[0]
        yardi_code = row[1]
        if not yardi_code:
            continue
        key = str(yardi_code).strip().upper()
        owned[key] = {"website_id": str(website_id).strip() if website_id else None}

    # --- Other Properties: richer metadata for enrichment ---
    ws_other = wb["Other Properties"]
    other_rows = list(ws_other.iter_rows(min_row=2, values_only=True))
    enrichment: dict[str, dict] = {}
    for row in other_rows:
        yardi_code = row[1]
        if not yardi_code:
            continue
        key = str(yardi_code).strip().upper()
        unit_count_raw = row[5]
        try:
            unit_count = int(unit_count_raw) if unit_count_raw is not None else None
        except (TypeError, ValueError):
            unit_count = None

        enrichment[key] = {
            "name": str(row[2]).strip() if row[2] else None,
            "state": str(row[11]).strip() if row[11] else None,
            "unit_count": unit_count,
        }

    # Merge: start with currently owned, enrich where possible
    for yardi_code, base in owned.items():
        extra = enrichment.get(yardi_code, {})
        session.merge(Property(
            yardi_code=yardi_code,
            website_id=base.get("website_id"),
            name=extra.get("name"),
            state=extra.get("state"),
            unit_count=extra.get("unit_count"),
        ))

    session.commit()
    wb.close()
    return len(owned)


def main() -> None:
    print("Initializing database schema...")
    init_db()

    with Session(engine) as session:
        print("Seeding GL codes...")
        gl_count = seed_gl_codes(session)
        print(f"  Inserted/replaced {gl_count} GL codes")

        print("Seeding properties...")
        prop_count = seed_properties(session)
        print(f"  Inserted/replaced {prop_count} properties")

    print("Seed complete.")


if __name__ == "__main__":
    main()
